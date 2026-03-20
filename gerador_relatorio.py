"""
Gerador de Relatorio Gerencial Mensal - BYM Gerenciamento
Baseado nos modelos: SETIN-Sampaio Viana e VITACON-Brigadeiro (Fevereiro-26)
"""

import io
import json
import os
from datetime import datetime, date
from typing import Optional, Union

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────── PALETA DE CORES ────────────────────────────────
COR_AZUL_ESCURO  = "1F3864"
COR_AZUL_MEDIO   = "2E75B6"
COR_AZUL_CLARO   = "BDD7EE"
COR_CINZA_CLARO  = "F2F2F2"
COR_BRANCO       = "FFFFFF"
COR_LARANJA      = "ED7D31"
COR_VERDE        = "70AD47"
COR_AMARELO      = "FFD966"
COR_VERMELHO     = "FF4B4B"
COR_VERDE_OK     = "C6EFCE"
COR_VERMELHO_BG  = "FFCCCC"
COR_TITULO_TEXTO = "FFFFFF"


# ─────────────────────────── HELPERS DE ESTILO ──────────────────────────────

def _fonte(negrito=False, tamanho=10, cor="000000", nome="Calibri"):
    return Font(name=nome, bold=negrito, size=tamanho, color=cor)

def _fill(cor_hex):
    return PatternFill("solid", fgColor=cor_hex)

def _alin(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _borda():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def _borda_bottom(cor="CCCCCC"):
    return Border(bottom=Side(style="thin", color=cor))

def _parse_data(valor) -> Optional[datetime]:
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime.combine(valor, datetime.min.time())
    try:
        return datetime.strptime(str(valor), "%Y-%m-%d")
    except ValueError:
        return None

def _cor_farol(farol) -> Optional[str]:
    return {1: COR_VERMELHO, 2: COR_AMARELO, 3: COR_VERDE}.get(farol)


# ─── células ────────────────────────────────────────────────────────────────

def _cell(ws, r, c, valor=None, negrito=False, tam=9, cor_f="000000",
          bg=None, h="left", wrap=False, borda=True, fmt=None):
    cell = ws.cell(row=r, column=c, value=valor)
    cell.font = _fonte(negrito, tam, cor_f)
    cell.alignment = _alin(h, "center", wrap)
    if bg:
        cell.fill = _fill(bg)
    if borda:
        cell.border = _borda()
    if fmt:
        cell.number_format = fmt
    return cell

def _header(ws, r, c1, c2, texto, bg=COR_AZUL_ESCURO, tam=11):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1, value=texto)
    c.font = _fonte(True, tam, COR_TITULO_TEXTO)
    c.fill = _fill(bg)
    c.alignment = _alin("center", "center")
    c.border = _borda()
    return c

def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _zebra(i):
    return COR_CINZA_CLARO if i % 2 == 0 else COR_BRANCO


# ─────────────────────────── CLASSE PRINCIPAL ───────────────────────────────

class GeradorRelatorio:

    def __init__(self, dados: Union[str, dict]):
        """Aceita caminho para JSON (str) ou dicionario de dados (dict)."""
        if isinstance(dados, str):
            with open(dados, encoding="utf-8") as f:
                self.dados = json.load(f)
        else:
            self.dados = dados
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def _periodo(self):
        r = self.dados["relatorio"]
        return f"{r['mes']}-{r['ano']}"

    def _topo(self, ws, titulo_secao: str):
        """Faixa de cabecalho padrao BYM em todas as abas."""
        ws.row_dimensions[1].height = 20
        ws.merge_cells("B1:I1")
        c = ws["B1"]
        c.value = "RELATORIO GERENCIAL MENSAL"
        c.font = _fonte(True, 12, COR_TITULO_TEXTO)
        c.fill = _fill(COR_AZUL_ESCURO)
        c.alignment = _alin("left", "center")

        ws.merge_cells("J1:L1")
        c2 = ws["J1"]
        c2.value = self._periodo()
        c2.font = _fonte(True, 12, COR_TITULO_TEXTO)
        c2.fill = _fill(COR_LARANJA)
        c2.alignment = _alin("center", "center")

        ws.row_dimensions[2].height = 18
        ws.merge_cells("B2:L2")
        c3 = ws["B2"]
        c3.value = titulo_secao
        c3.font = _fonte(True, 11, COR_TITULO_TEXTO)
        c3.fill = _fill(COR_AZUL_MEDIO)
        c3.alignment = _alin("center", "center")

    # ── CAPA ─────────────────────────────────────────────────────────────────

    def _criar_capa(self):
        ws = self.wb.create_sheet("CAPA")
        ws.sheet_view.showGridLines = False
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12

        for row in range(3, 30):
            for col in range(2, 14):
                ws.cell(row=row, column=col).fill = _fill(COR_AZUL_ESCURO)

        ws.row_dimensions[3].height = 8
        for col in range(2, 14):
            ws.cell(row=3, column=col).fill = _fill(COR_LARANJA)

        p = self.dados["projeto"]

        ws.row_dimensions[6].height = 22
        ws.merge_cells("B6:M6")
        c = ws["B6"]
        c.value = f"  {p.get('incorporadora', '')}  |  {p.get('construtora', '')}"
        c.font = _fonte(True, 13, COR_BRANCO)
        c.fill = _fill(COR_AZUL_ESCURO)
        c.alignment = _alin("center", "center")

        ws.row_dimensions[10].height = 36
        ws.merge_cells("B10:M10")
        c2 = ws["B10"]
        c2.value = p["nome"]
        c2.font = _fonte(True, 22, COR_BRANCO)
        c2.fill = _fill(COR_AZUL_ESCURO)
        c2.alignment = _alin("center", "center")

        ws.row_dimensions[14].height = 28
        ws.merge_cells("B14:M14")
        c3 = ws["B14"]
        c3.value = "RELATORIO GERENCIAL MENSAL"
        c3.font = _fonte(True, 16, COR_BRANCO)
        c3.fill = _fill(COR_AZUL_ESCURO)
        c3.alignment = _alin("center", "center")

        ws.row_dimensions[18].height = 28
        ws.merge_cells("B18:M18")
        c4 = ws["B18"]
        c4.value = self._periodo()
        c4.font = _fonte(True, 20, COR_TITULO_TEXTO)
        c4.fill = _fill(COR_LARANJA)
        c4.alignment = _alin("center", "center")

        ws.row_dimensions[29].height = 8
        for col in range(2, 14):
            ws.cell(row=29, column=col).fill = _fill(COR_LARANJA)

        for i, (label, valor) in enumerate([
            ("Incorporadora:", p["incorporadora"]),
            ("Construtora:", p["construtora"]),
            ("Gerenciadora:", p["gerenciadora"]),
        ]):
            r = 31 + i * 2
            ws.merge_cells(f"B{r}:D{r}")
            c = ws[f"B{r}"]
            c.value = label
            c.font = _fonte(True, 11)
            c.alignment = _alin("right", "center")
            ws.merge_cells(f"E{r}:H{r}")
            c2 = ws[f"E{r}"]
            c2.value = valor
            c2.font = _fonte(False, 11)
            c2.alignment = _alin("left", "center")

    # ── SUMARIO ──────────────────────────────────────────────────────────────

    def _criar_sumario(self):
        ws = self.wb.create_sheet("Sumario")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["B"].width = 65
        self._topo(ws, "SUMARIO")

        itens = [
            ("1. Dados e Caracteristicas do Empreendimento", 0),
            ("2. Controle de Prazo", 0),
            ("   2.1. Prazos Contratuais x Tendencia", 1),
            ("   2.2. Avanco Fisico no Periodo", 1),
            ("   2.3. Datas Marco Contratual", 1),
            ("3. Controle de Custo", 0),
            ("   3.1. Avanco de Custo no Periodo (Aporte)", 1),
            ("   3.2. Fluxo de Caixa", 1),
            ("   3.3. Analise Financeira por Centro de Custo", 1),
            ("4. Graficos Resumo", 0),
            ("5. Datas Marco Prototipo", 0),
            ("6. Histograma de Mao de Obra", 0),
            ("   6.1. Mao de Obra Construtora", 1),
            ("   6.2. Mao de Obra Terceirizada", 1),
            ("7. Evolucao do Cronograma", 0),
            ("   7.1. Espelho do Cronograma Mensal", 1),
            ("   7.2. Espelho do Cronograma Geral", 1),
            ("8. Farol da Lista de Metas do Periodo", 0),
            ("   8.1. Desvios dos Grupos", 1),
            ("9. Lista de Metas Para o Proximo Periodo", 0),
            ("10. Controle de Processos p/ Obtencao de Habite-se", 0),
            ("   10.1. Consideracoes de prazo", 1),
            ("   10.2. Lista de Documentos Iniciais", 1),
            ("   10.4. Cronograma de Legalizacao e Concessionarias", 1),
            ("11. Grafico Fisico s/ Estoque", 0),
            ("12. Grafico Fisico c/ Estoque X Aporte", 0),
            ("13. Custo Realizado - Tabela de Aportes", 0),
            ("14. Gerenciamento de Contratacoes", 0),
            ("15. Controle de Mapas de Contratacoes", 0),
            ("16. Cronograma de Suprimentos", 0),
            ("17. Relatorio Fotografico", 0),
        ]

        for i, (texto, nivel) in enumerate(itens):
            r = 4 + i
            ws.row_dimensions[r].height = 17
            ws.merge_cells(f"B{r}:M{r}")
            c = ws[f"B{r}"]
            c.value = texto
            c.font = _fonte(negrito=(nivel == 0), tamanho=10,
                           cor=COR_AZUL_ESCURO if nivel == 0 else "000000")
            c.alignment = _alin("left", "center")
            c.border = _borda_bottom()
            if nivel == 0:
                c.fill = _fill(COR_AZUL_CLARO)

    # ── 1-4) RESUMO E INDICADORES ────────────────────────────────────────────

    def _criar_resumo_indicadores(self):
        ws = self.wb.create_sheet("1-4) Resumo e Indicadores")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 26, "C": 3, "D": 26, "E": 15,
                             "F": 15, "G": 14, "H": 14, "I": 14})
        self._topo(ws, "RESUMO E INDICADORES")

        p = self.dados["projeto"]
        pr = self.dados["controle_prazo"]
        av = self.dados["avanco_fisico"]

        # -- Secao 1: Dados do Empreendimento --
        r = 4
        _header(ws, r, 2, 9, "1. Dados e Caracteristicas do Empreendimento",
                COR_AZUL_MEDIO, 10)

        esq = [
            ("Incorporadora:", p["incorporadora"]),
            ("Construtora:", p["construtora"]),
            ("Gerenciadora:", p["gerenciadora"]),
            ("Padrao:", p["padrao"]),
            ("Endereco:", p["endereco"]),
            ("Area do Terreno:", f"{p['area_terreno']:,.2f} m2 ({p['area_terreno_referencia']})"),
            ("Area Construida:", f"{p['area_construida']:,.2f} m2 ({p['area_construida_referencia']})"),
            ("Area Privativa:", f"{p['area_privativa']:,.2f} m2 ({p['area_privativa_referencia']})"),
        ]
        dir_ = [
            ("Torres:", str(p["torres"])),
            ("N de Fases:", str(p["num_fases"])),
            ("Unidades:", str(p["unidades"])),
            ("Subsolos/vagas:", p["subsolos_vagas"]),
            ("N de pavimentos:", p["num_pavimentos"]),
            ("Tipo de contencao:", p["tipo_contencao"]),
            ("Tipo de fundacao:", p["tipo_fundacao"]),
            ("Tipo de estrutura:", p["tipo_estrutura"]),
        ]
        for i, ((le, ve), (ld, vd)) in enumerate(zip(esq, dir_)):
            row = r + 1 + i
            ws.row_dimensions[row].height = 16
            bg = _zebra(i)
            _cell(ws, row, 2, le, negrito=True, tam=9, bg=bg)
            ws.merge_cells(start_row=row, start_column=3,
                          end_row=row, end_column=5)
            _cell(ws, row, 3, ve, tam=9, bg=bg)
            _cell(ws, row, 6, ld, negrito=True, tam=9, bg=bg)
            ws.merge_cells(start_row=row, start_column=7,
                          end_row=row, end_column=9)
            _cell(ws, row, 7, vd, tam=9, bg=bg)

        # -- Secao 2: Controle de Prazo --
        r = 14
        _header(ws, r, 2, 9, "2. Controle de Prazo", COR_AZUL_MEDIO, 10)

        r += 1
        _header(ws, r, 2, 9, "2.1. Prazos Contratuais x Tendencia",
                COR_AZUL_CLARO, 9)
        r += 1
        for col, txt in [(5, "Contratual"), (6, "Tendencia"), (7, "Desvio"), (8, "Unid.")]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")

        linhas_prazo = [
            ("Data de inicio Contratual:", pr["data_inicio_contratual"],
             pr["data_inicio_tendencia"], 0, "DIAS"),
            ("Data de Termino Contratual:", pr["data_termino_contratual"],
             pr["data_termino_tendencia"], pr.get("desvio_dias", 0), "DIAS"),
            ("Qtde Meses:", pr["qtde_meses_contratual"],
             pr["qtde_meses_tendencia"],
             round(pr["qtde_meses_tendencia"] - pr["qtde_meses_contratual"], 1), "MESES"),
        ]
        for i, (label, contr, tend, desv, unid) in enumerate(linhas_prazo):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            ws.merge_cells(start_row=r, start_column=2,
                          end_row=r, end_column=4)
            _cell(ws, r, 2, label, negrito=True, tam=9, bg=bg)
            vc = _parse_data(contr) if isinstance(contr, str) else contr
            vt = _parse_data(tend) if isinstance(tend, str) else tend
            fmt = "DD/MM/YYYY" if isinstance(vc, datetime) else None
            _cell(ws, r, 5, vc, tam=9, bg=bg, h="center", fmt=fmt)
            _cell(ws, r, 6, vt, tam=9, bg=bg, h="center", fmt=fmt)
            bg_d = COR_VERDE_OK if (desv == 0) else COR_VERMELHO_BG
            _cell(ws, r, 7, desv, tam=9, bg=bg_d, h="center")
            _cell(ws, r, 8, unid, tam=8, bg=bg, h="center")

        # -- Datas Marco Contratual --
        r += 2
        _header(ws, r, 2, 9, "2.3. Datas Marco Contratual",
                COR_AZUL_MEDIO, 10)
        r += 1
        for col, txt in [(2, "ATIVIDADE"), (3, "Baseline\nMeses"),
                         (4, "Baseline\nData"), (5, "Previsto\nMeses"),
                         (6, "Previsto\nData"), (7, "Desvio\n(dias)"), (8, "Farol")]:
            _cell(ws, r, col, txt, negrito=True, tam=8,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center", wrap=True)

        marcos = pr.get("datas_marco_contratual", [])
        for i, m in enumerate(marcos):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            _cell(ws, r, 2, m["atividade"], tam=9, bg=bg)
            _cell(ws, r, 3, m["baseline_meses"], tam=9, bg=bg, h="center")
            _cell(ws, r, 4, _parse_data(m["baseline_data"]),
                  tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
            _cell(ws, r, 5, m["previsto_meses"], tam=9, bg=bg, h="center")
            _cell(ws, r, 6, _parse_data(m["previsto_data"]),
                  tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
            dd = m["desvio_dias"]
            bg_d = COR_VERDE_OK if dd == 0 else (COR_AMARELO if dd <= 30 else COR_VERMELHO_BG)
            _cell(ws, r, 7, dd, tam=9, bg=bg_d, h="center")
            cf = _cell(ws, r, 8, "", tam=9, h="center")
            if m.get("farol"):
                cf.fill = _fill(_cor_farol(m["farol"]))

        # -- Avanco Fisico --
        r += 2
        _header(ws, r, 2, 9, "2.2. Avanco Fisico no Periodo",
                COR_AZUL_MEDIO, 10)
        r += 1
        for col, txt in [(2, "Indicador"), (3, "Mes"), (4, "Acumulado")]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")

        linhas_av = [
            ("Previsto Curva Contratual",
             av.get("previsto_curva_contratual_mes"),
             av.get("previsto_curva_contratual_acum")),
            ("Realizado Medido",
             av.get("realizado_medido_mes"),
             av.get("realizado_medido_acum")),
            ("Desvio",
             av.get("desvio_mes"),
             av.get("desvio_acumulado")),
            ("IPF Contratual", av.get("ipf_contratual"), None),
            ("Meta BYM",
             av.get("meta_bym_mes"),
             av.get("meta_bym_acum")),
            ("IPF Meta BYM", av.get("ipf_meta_bym"), None),
        ]
        for i, (label, vmes, vacum) in enumerate(linhas_av):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            _cell(ws, r, 2, label, negrito=True, tam=9, bg=bg)
            is_desvio = "Desvio" in label
            bg_m = (COR_VERDE_OK if (vmes or 0) >= 0 else COR_VERMELHO_BG) if is_desvio else bg
            bg_a = (COR_VERDE_OK if (vacum or 0) >= 0 else COR_VERMELHO_BG) if is_desvio else bg
            _cell(ws, r, 3, vmes, tam=9, bg=bg_m, h="center",
                  fmt="0.00%" if vmes is not None else None)
            _cell(ws, r, 4, vacum, tam=9, bg=bg_a, h="center",
                  fmt="0.00%" if vacum is not None else None)

        # -- Qtde servicos --
        r += 2
        _header(ws, r, 2, 9, "Servicos no Periodo", COR_AZUL_MEDIO, 9)
        r += 1
        for col, txt in [(2, "Qtde Previstos"), (3, "Qtde Executados"), (4, "% Atingido")]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")
        r += 1
        _cell(ws, r, 2, av.get("qtde_servicos_previstos_mes"), tam=10,
              negrito=True, bg=COR_AZUL_CLARO, h="center")
        _cell(ws, r, 3, av.get("qtde_servicos_executados_mes"), tam=10,
              negrito=True, bg=COR_AZUL_CLARO, h="center")
        _cell(ws, r, 4, av.get("percentual_atingido"), tam=10,
              negrito=True, bg=COR_AZUL_CLARO, h="center", fmt="0.00%")

    # ── 3.2) FLUXO DE CAIXA ──────────────────────────────────────────────────

    def _criar_fluxo_caixa(self):
        ws = self.wb.create_sheet("3.2) Fluxo de Caixa")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 4, "C": 10, "D": 12, "E": 12,
                             "F": 10, "G": 10, "H": 14, "I": 10, "J": 10, "K": 9})
        self._topo(ws, "GERENCIAMENTO DO CUSTO")

        fc = self.dados.get("fluxo_caixa", {})
        r = 4
        _header(ws, r, 2, 11, "3.2. Fluxo de Caixa - Previsto x Realizado",
                COR_AZUL_MEDIO, 10)
        r += 1

        # Info contrato
        _cell(ws, r, 2, "Valor Total do Contrato (R$):", negrito=True, tam=9,
              bg=COR_AZUL_CLARO, borda=False)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        _cell(ws, r, 3, fc.get("valor_total_contrato"), negrito=True, tam=10,
              bg=COR_AZUL_CLARO, h="center", fmt='R$ #,##0.00')
        _cell(ws, r, 5, f"INCC BASE: {fc.get('incc_base_referencia', '')}",
              tam=8, bg=COR_AZUL_CLARO, h="center")
        _cell(ws, r, 6, fc.get("incc_base"), tam=9, bg=COR_AZUL_CLARO,
              h="center", fmt="#,##0.000")
        r += 2

        # Grupo "Previsto"
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        _cell(ws, r, 4, "Previsto Competencia", negrito=True, tam=9,
              bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=10)
        _cell(ws, r, 7, "Realizado Caixa", negrito=True, tam=9,
              bg=COR_LARANJA, cor_f=COR_TITULO_TEXTO, h="center")
        r += 1

        cabecalhos = [
            (2, "#"), (3, "Mes"), (4, "Valor (R$)"), (5, "INCC"), (6, "%"),
            (7, "Valor (R$)"), (8, "INCC"), (9, "%"), (10, "Desvio %"), (11, "Status"),
        ]
        for col, txt in cabecalhos:
            _cell(ws, r, col, txt, negrito=True, tam=8,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center", wrap=True)

        for i, m in enumerate(fc.get("meses", [])):
            r += 1
            ws.row_dimensions[r].height = 15
            bg = _zebra(i)
            realizado = m.get("realizado_valor") is not None
            desvio = m.get("desvio_pct")

            _cell(ws, r, 2, m["mes"], tam=8, bg=bg, h="center")
            _cell(ws, r, 3, _parse_data(m["data"]), tam=8, bg=bg,
                  h="center", fmt="MMM/YY")
            _cell(ws, r, 4, m.get("previsto_valor"), tam=8, bg=bg,
                  h="center", fmt='#,##0.00')
            _cell(ws, r, 5, m.get("previsto_incc"), tam=8, bg=bg,
                  h="center", fmt='#,##0.00')
            _cell(ws, r, 6, m.get("previsto_pct"), tam=8, bg=bg,
                  h="center", fmt="0.00%")
            _cell(ws, r, 7, m.get("realizado_valor"), tam=8, bg=bg,
                  h="center", fmt='#,##0.00')
            _cell(ws, r, 8, m.get("realizado_incc"), tam=8, bg=bg,
                  h="center", fmt='#,##0.00')
            _cell(ws, r, 9, m.get("realizado_pct"), tam=8, bg=bg,
                  h="center", fmt="0.00%")

            if realizado and desvio is not None:
                bg_d = COR_VERDE_OK if desvio >= 0 else COR_VERMELHO_BG
                _cell(ws, r, 10, desvio, tam=8, bg=bg_d, h="center", fmt="0.00%")
            else:
                _cell(ws, r, 10, None, tam=8, bg=bg, h="center")

            status = "Realizado" if realizado else "Previsto"
            bg_s = COR_VERDE_OK if realizado else COR_CINZA_CLARO
            _cell(ws, r, 11, status, tam=8, bg=bg_s, h="center")

    # ── 3.3) ANALISE FINANCEIRA ───────────────────────────────────────────────

    def _criar_analise_financeira(self):
        ws = self.wb.create_sheet("3.3) Analise Financeira")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 4, "C": 12, "D": 40, "E": 14,
                             "F": 12, "G": 10})
        self._topo(ws, "GERENCIAMENTO DO CUSTO")

        af = self.dados.get("analise_financeira", {})
        r = 4
        _header(ws, r, 2, 7,
                "3.3. Analise Financeira por Centro de Custo",
                COR_AZUL_MEDIO, 10)
        r += 1

        cabecalhos = [
            (2, "Nivel"), (3, "ID"), (4, "Descricao"),
            (5, "Orcamento R$"), (6, "INCC"), (7, "Peso %"),
        ]
        for col, txt in cabecalhos:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")

        linha = 0
        for item in af.get("itens", []):
            r += 1
            ws.row_dimensions[r].height = 17
            _cell(ws, r, 2, "Grupo", tam=8, bg=COR_AZUL_MEDIO,
                  cor_f=COR_TITULO_TEXTO, h="center")
            _cell(ws, r, 3, item["id"], tam=9, bg=COR_AZUL_MEDIO,
                  cor_f=COR_TITULO_TEXTO)
            _cell(ws, r, 4, item["descricao"], negrito=True, tam=9,
                  bg=COR_AZUL_MEDIO, cor_f=COR_TITULO_TEXTO)
            _cell(ws, r, 5, item["orcamento_rs"], tam=9,
                  bg=COR_AZUL_MEDIO, cor_f=COR_TITULO_TEXTO,
                  h="center", fmt='#,##0.00')
            _cell(ws, r, 6, item["orcamento_incc"], tam=9,
                  bg=COR_AZUL_MEDIO, cor_f=COR_TITULO_TEXTO,
                  h="center", fmt='#,##0.00')
            _cell(ws, r, 7, item["peso_pct"], tam=9,
                  bg=COR_AZUL_MEDIO, cor_f=COR_TITULO_TEXTO,
                  h="center", fmt="0.000%")

            for sub in item.get("subitens", []):
                r += 1
                ws.row_dimensions[r].height = 15
                bg = _zebra(linha)
                linha += 1
                _cell(ws, r, 2, "Item", tam=8, bg=bg, h="center")
                _cell(ws, r, 3, sub["id"], tam=8, bg=bg)
                _cell(ws, r, 4, f"   {sub['descricao']}", tam=9, bg=bg)
                _cell(ws, r, 5, sub["orcamento_rs"], tam=9, bg=bg,
                      h="center", fmt='#,##0.00')
                _cell(ws, r, 6, sub["orcamento_incc"], tam=9, bg=bg,
                      h="center", fmt='#,##0.00')
                _cell(ws, r, 7, sub["peso_pct"], tam=9, bg=bg,
                      h="center", fmt="0.000%")

        # Total
        r += 1
        total = sum(it["orcamento_rs"] for it in af.get("itens", []))
        _header(ws, r, 2, 4, "TOTAL ORCAMENTO", COR_AZUL_ESCURO, 9)
        _cell(ws, r, 5, total, negrito=True, tam=10,
              bg=COR_AZUL_CLARO, h="center", fmt='#,##0.00')

    # ── 5-6) PROTOTIPO E HISTOGRAMA ──────────────────────────────────────────

    def _criar_prototipo_histograma(self):
        ws = self.wb.create_sheet("5-6) Prototipo e Histograma")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 24, "C": 24, "D": 13, "E": 13,
                             "F": 13, "G": 10, "H": 8, "I": 10,
                             "J": 10, "K": 10})
        self._topo(ws, "GERENCIAMENTO DO PRAZO")

        # -- 5. Datas Marco Prototipo --
        r = 4
        _header(ws, r, 2, 9, "5. Datas Marco Prototipo", COR_AZUL_MEDIO, 10)
        r += 1
        for col, txt in [
            (2, "ATIVIDADE"), (3, "LOCAL"), (4, "DATA CONTRATUAL"),
            (5, "DATA PREVISTA (ENG)"), (6, "DATA PREVISTA (DIR)"),
            (7, "Desvio (dias)"), (8, "Farol"),
        ]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        for i, m in enumerate(self.dados.get("datas_marco_prototipo", [])):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            _cell(ws, r, 2, m["atividade"], tam=9, bg=bg)
            _cell(ws, r, 3, m["local"], tam=9, bg=bg)
            _cell(ws, r, 4, _parse_data(m["data_contratual"]),
                  tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
            _cell(ws, r, 5, _parse_data(m["data_prevista_eng"]),
                  tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
            _cell(ws, r, 6, _parse_data(m["data_prevista_dir"]),
                  tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
            dd = m["desvio_dias"]
            bg_d = COR_VERDE_OK if dd == 0 else (COR_AMARELO if dd <= 30 else COR_VERMELHO_BG)
            _cell(ws, r, 7, dd, tam=9, bg=bg_d, h="center")
            cf = _cell(ws, r, 8, "", tam=9, h="center")
            if m.get("farol"):
                cf.fill = _fill(_cor_farol(m["farol"]))

        # -- 6. Histograma de Mao de Obra --
        r += 2
        _header(ws, r, 2, 11, "6. Histograma de Mao de Obra - Terceirizada",
                COR_AZUL_MEDIO, 10)
        r += 1

        hist = self.dados["histograma_mao_obra"]
        d_ref = _parse_data(hist["mes_referencia"])
        meses_nomes = ["Jan","Fev","Mar","Abr","Mai","Jun",
                       "Jul","Ago","Set","Out","Nov","Dez"]
        if d_ref:
            m2 = meses_nomes[(d_ref.month - 3) % 12]
            m1 = meses_nomes[(d_ref.month - 2) % 12]
        else:
            m2, m1 = "M-2", "M-1"

        cabs = ["Empresa", "Servico", m2, m1, "Mes Atual",
                "Sem 1", "Sem 2", "Sem 3", "Sem 4", "TOTAL"]
        for j, txt in enumerate(cabs):
            _cell(ws, r, j + 2, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        for i, emp in enumerate(hist["empresas"]):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            vals = [emp["nome"], emp["servico"],
                    emp["mes_anterior2"], emp["mes_anterior1"],
                    emp["mes_atual"], emp.get("semana1"),
                    emp.get("semana2"), emp.get("semana3"),
                    emp.get("semana4"), emp["total"]]
            for j, v in enumerate(vals):
                _cell(ws, r, j + 2, v, tam=9, bg=bg,
                      h="center" if j >= 2 else "left")

    # ── 8) FAROL DE METAS ────────────────────────────────────────────────────

    def _criar_farol_metas(self):
        ws = self.wb.create_sheet("8) Farol de Metas")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 4, "C": 34, "D": 10, "E": 9,
                             "F": 13, "G": 13, "H": 13, "I": 7, "J": 22})
        self._topo(ws, "GERENCIAMENTO DO PRAZO")

        av = self.dados["avanco_fisico"]
        r = 4
        _header(ws, r, 2, 10, "8. Farol da Lista de Metas do Periodo",
                COR_AZUL_MEDIO, 10)
        r += 1

        for label, valor, fmt in [
            ("Qtde Servicos Previstos:", av["qtde_servicos_previstos_mes"], None),
            ("Qtde Servicos Executados:", av["qtde_servicos_executados_mes"], None),
            ("% Atingido da Meta:", av["percentual_atingido"], "0.00%"),
        ]:
            ws.row_dimensions[r].height = 17
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            _cell(ws, r, 2, label, negrito=True, tam=10, bg=COR_AZUL_CLARO)
            _cell(ws, r, 6, valor, negrito=True, tam=10,
                  bg=COR_AZUL_CLARO, h="center", fmt=fmt)
            r += 1

        # Legenda
        r += 1
        for cf_hex, txt in [(COR_VERDE, "Executado no prazo"),
                            (COR_AMARELO, "Parcialmente executado"),
                            (COR_VERMELHO, "Nao executado")]:
            c = ws.cell(row=r, column=2)
            c.fill = _fill(cf_hex)
            c.border = _borda()
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            _cell(ws, r, 3, txt, tam=9)
            r += 1

        # Cabecalhos
        r += 1
        for col, txt in [(2, "#"), (3, "Nome da Tarefa"), (4, "% Conclui."),
                         (5, "Duracao"), (6, "Inicio"), (7, "Termino"),
                         (8, "Termino Real"), (9, "Farol"), (10, "Observacoes")]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")

        num = 1
        ultima_torre = None
        for grupo in self.dados["farol_metas"]:
            torre = grupo.get("torre", "")
            if torre and torre != ultima_torre:
                r += 1
                ws.merge_cells(start_row=r, start_column=2,
                               end_row=r, end_column=10)
                c = ws.cell(row=r, column=2, value=torre)
                c.font = _fonte(True, 10, COR_TITULO_TEXTO)
                c.fill = _fill(COR_AZUL_ESCURO)
                c.alignment = _alin("center", "center")
                c.border = _borda()
                ultima_torre = torre

            r += 1
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=10)
            c = ws.cell(row=r, column=2,
                        value=f"   {grupo['categoria']}")
            c.font = _fonte(True, 9, COR_TITULO_TEXTO)
            c.fill = _fill(COR_AZUL_MEDIO)
            c.alignment = _alin("left", "center")
            c.border = _borda()

            for item in grupo["itens"]:
                r += 1
                ws.row_dimensions[r].height = 15
                bg = _zebra(num)
                _cell(ws, r, 2, num, tam=9, bg=bg, h="center")
                _cell(ws, r, 3, f"      {item['nome']}", tam=9, bg=bg)
                _cell(ws, r, 4,
                      (item["percentual"] / 100 if item["percentual"] else 0),
                      tam=9, bg=bg, h="center", fmt="0%")
                _cell(ws, r, 5, item["duracao"], tam=9, bg=bg, h="center")
                _cell(ws, r, 6, _parse_data(item["inicio"]),
                      tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
                _cell(ws, r, 7, _parse_data(item["termino"]),
                      tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
                _cell(ws, r, 8, _parse_data(item.get("termino_real")),
                      tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
                cf = _cell(ws, r, 9, "", tam=9, h="center")
                if item.get("farol"):
                    cf.fill = _fill(_cor_farol(item["farol"]))
                _cell(ws, r, 10, item.get("observacoes", ""),
                      tam=8, bg=bg, wrap=True)
                num += 1

    # ── 9) METAS PROXIMO MES ─────────────────────────────────────────────────

    def _criar_metas_proximo_mes(self):
        ws = self.wb.create_sheet("9) Metas Proximo Mes")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 34, "C": 10, "D": 10,
                             "E": 13, "F": 13, "G": 13, "H": 13})
        self._topo(ws, "GERENCIAMENTO DO PRAZO")

        r = 4
        _header(ws, r, 2, 8,
                "9. Lista de Metas Para o Proximo Periodo",
                COR_AZUL_MEDIO, 10)
        r += 1

        for col, txt in [(2, "Nome da Tarefa"), (3, "% Conclui."),
                         (4, "Duracao"), (5, "Inicio"), (6, "Termino"),
                         (7, "Baseline Inicio"), (8, "Baseline Termino")]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        ultima_torre = None
        for grupo in self.dados["metas_proximo_mes"]:
            torre = grupo.get("torre", "")
            if torre and torre != ultima_torre:
                r += 1
                ws.merge_cells(start_row=r, start_column=2,
                               end_row=r, end_column=8)
                c = ws.cell(row=r, column=2, value=torre)
                c.font = _fonte(True, 10, COR_TITULO_TEXTO)
                c.fill = _fill(COR_AZUL_ESCURO)
                c.alignment = _alin("center", "center")
                c.border = _borda()
                ultima_torre = torre

            r += 1
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=8)
            c = ws.cell(row=r, column=2,
                        value=f"   {grupo['categoria']}")
            c.font = _fonte(True, 9, COR_TITULO_TEXTO)
            c.fill = _fill(COR_AZUL_MEDIO)
            c.alignment = _alin("left", "center")
            c.border = _borda()

            for i, item in enumerate(grupo["itens"]):
                r += 1
                bg = _zebra(i)
                _cell(ws, r, 2, f"      {item['nome']}", tam=9, bg=bg)
                _cell(ws, r, 3,
                      (item["percentual"] / 100 if item["percentual"] else 0),
                      tam=9, bg=bg, h="center", fmt="0%")
                _cell(ws, r, 4, item["duracao"], tam=9, bg=bg, h="center")
                _cell(ws, r, 5, _parse_data(item["inicio"]),
                      tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
                _cell(ws, r, 6, _parse_data(item["termino"]),
                      tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
                _cell(ws, r, 7,
                      _parse_data(item.get("baseline_inicio")),
                      tam=9, bg=COR_CINZA_CLARO, h="center", fmt="DD/MM/YYYY")
                _cell(ws, r, 8,
                      _parse_data(item.get("baseline_termino")),
                      tam=9, bg=COR_CINZA_CLARO, h="center", fmt="DD/MM/YYYY")

    # ── 10) LEGALIZACAO ───────────────────────────────────────────────────────

    def _criar_legalizacao(self):
        ws = self.wb.create_sheet("10) Legalizacao")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 48, "C": 14, "D": 14,
                             "E": 10, "F": 8})
        self._topo(ws, "CRONOGRAMA E STATUS LEGALIZACOES")

        leg = self.dados["legalizacao"]
        r = 4
        _header(ws, r, 2, 7,
                "10. Controle de Processos p/ Obtencao de Habite-se",
                COR_AZUL_MEDIO, 10)
        r += 1
        _header(ws, r, 2, 7, "10.1. Consideracoes de prazo",
                COR_AZUL_CLARO, 9)
        r += 1

        for label, data in [
            ("I - Inicio da obra", leg["inicio_obra"]),
            ("H - Mes Habite-se da obra", leg["mes_habitese"]),
            ("T - Mes termino da obra", leg["mes_termino"]),
        ]:
            ws.row_dimensions[r].height = 16
            ws.merge_cells(start_row=r, start_column=2,
                          end_row=r, end_column=3)
            _cell(ws, r, 2, label, negrito=True, tam=9)
            _cell(ws, r, 4, _parse_data(data),
                  tam=9, h="center", fmt="DD/MM/YYYY")
            r += 1

        r += 1
        _header(ws, r, 2, 7, "10.2. Lista de Documentos Iniciais",
                COR_AZUL_MEDIO, 10)
        r += 1
        for col, txt in [(2, "Documento"), (3, "Data Disponibilizacao"),
                         (4, "Data de Validade"), (5, "Prazo"), (6, "Status")]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        for i, doc in enumerate(leg["documentos_iniciais"]):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            status = doc["status"]
            bg_s = COR_VERDE_OK if status == "OK" else COR_AMARELO
            _cell(ws, r, 2, doc["documento"], tam=9, bg=bg, wrap=True)
            _cell(ws, r, 3, _parse_data(doc["disponibilizado"]),
                  tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
            val_v = (_parse_data(doc["validade"])
                     if doc["validade"] not in ("N/A", "NA", None)
                     else doc["validade"])
            _cell(ws, r, 4, val_v, tam=9, bg=bg, h="center",
                  fmt="DD/MM/YYYY" if isinstance(val_v, datetime) else None)
            _cell(ws, r, 5, doc["prazo"], tam=9, bg=bg, h="center")
            _cell(ws, r, 6, status, negrito=True, tam=9,
                  bg=bg_s, h="center")

    # ── 13) TABELA DE APORTE ─────────────────────────────────────────────────

    def _criar_tabela_aporte(self):
        ws = self.wb.create_sheet("13) Tabela de Aporte")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 4, "C": 10, "D": 9, "E": 9,
                             "F": 12, "G": 16, "H": 10, "I": 9,
                             "J": 16, "K": 10, "L": 9, "M": 10, "N": 9})
        self._topo(ws, "GERENCIAMENTO DO CUSTO")

        r = 4
        _header(ws, r, 2, 14, "13. Custo Realizado", COR_AZUL_MEDIO, 10)
        r += 1
        _header(ws, r, 2, 14,
                "13.1. Tabela de Aportes - Previsto x Realizado",
                COR_AZUL_CLARO, 9)
        r += 1

        # linha de grupo
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        _cell(ws, r, 4, "Previsto", negrito=True, tam=9,
              bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=9)
        _cell(ws, r, 6, "Realizado", negrito=True, tam=9,
              bg=COR_LARANJA, cor_f=COR_TITULO_TEXTO, h="center")
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
        _cell(ws, r, 10, "Acumulado", negrito=True, tam=9,
              bg=COR_AZUL_MEDIO, cor_f=COR_TITULO_TEXTO, h="center")
        r += 1

        for col, txt in [(2, "#"), (3, "Mes"), (4, "% Mes"), (5, "% Acum"),
                         (6, "Data Aporte"), (7, "Valor (R$)"),
                         (8, "INCC"), (9, "%"),
                         (10, "Valor (R$)"), (11, "INCC"), (12, "%"),
                         (13, "INCC Utilizado"), (14, "Desvio %")]:
            _cell(ws, r, col, txt, negrito=True, tam=8,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        for i, ap in enumerate(self.dados["tabela_aporte"]):
            r += 1
            ws.row_dimensions[r].height = 15
            bg = _zebra(i)
            d = ap["desvio_pct"]
            bg_d = COR_VERDE_OK if d >= 0 else COR_VERMELHO_BG

            for col, val, fmt_, bg_ in [
                (2, ap["mes"], None, bg),
                (3, _parse_data(ap["data"]), "MMM/YY", bg),
                (4, ap["previsto_pct_mes"], "0.00%", bg),
                (5, ap["previsto_pct_acum"], "0.00%", bg),
                (6, _parse_data(ap["data_aporte"]), "DD/MM/YYYY", bg),
                (7, ap["realizado_valor"], "#,##0.00", bg),
                (8, ap["realizado_incc"], "#,##0.00", bg),
                (9, ap["realizado_pct"], "0.00%", bg),
                (10, ap["acum_valor"], "#,##0.00", bg),
                (11, ap["acum_incc"], "#,##0.00", bg),
                (12, ap["acum_pct"], "0.00%", bg),
                (13, ap["incc_utilizado"], "#,##0.000", bg),
                (14, d, "0.00%", bg_d),
            ]:
                _cell(ws, r, col, val, tam=8, bg=bg_, h="center", fmt=fmt_)

        r += 1
        ult = self.dados["tabela_aporte"][-1]
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        _cell(ws, r, 2, "TOTAL ACUMULADO", negrito=True, tam=9,
              bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")
        _cell(ws, r, 10, ult["acum_valor"], negrito=True, tam=9,
              bg=COR_AZUL_CLARO, h="center", fmt="#,##0.00")
        _cell(ws, r, 12, ult["acum_pct"], negrito=True, tam=9,
              bg=COR_AZUL_CLARO, h="center", fmt="0.00%")

    # ── 14) GERENCIAMENTO DE CONTRATACOES ────────────────────────────────────

    def _criar_gerenciamento_contratacoes(self):
        ws = self.wb.create_sheet("14) Gerenc. Contratacoes")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 4, "C": 5, "D": 30,
                             "E": 14, "F": 14, "G": 8, "H": 24})
        self._topo(ws, "GERENCIAMENTO DE CONTRATACOES")

        r = 4
        _header(ws, r, 2, 8,
                "14. Cronograma de Contratacoes",
                COR_AZUL_MEDIO, 10)
        r += 1

        for col, txt in [(2, "#"), (3, "ID"), (4, "Servico"),
                         (5, "Data Prevista"), (6, "Data Contratacao"),
                         (7, "Farol"), (8, "Fornecedor")]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        for i, con in enumerate(self.dados.get("gerenciamento_contratacoes", [])):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            _cell(ws, r, 2, i + 1, tam=9, bg=bg, h="center")
            _cell(ws, r, 3, con["id"], tam=9, bg=bg, h="center")
            _cell(ws, r, 4, con["servico"], tam=9, bg=bg)
            _cell(ws, r, 5, _parse_data(con["data_prevista"]),
                  tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
            _cell(ws, r, 6, _parse_data(con.get("data_contratacao")),
                  tam=9, bg=bg, h="center", fmt="DD/MM/YYYY")
            cf = _cell(ws, r, 7, "", tam=9, h="center")
            if con.get("farol"):
                cf.fill = _fill(_cor_farol(con["farol"]))
            _cell(ws, r, 8, con.get("fornecedor", ""), tam=9, bg=bg)

    # ── 15) CONTROLE DE MAPAS DE CONTRATACOES ────────────────────────────────

    def _criar_controle_mapas_contratacoes(self):
        ws = self.wb.create_sheet("15) Mapas de Contratacoes")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {"B": 4, "C": 8, "D": 28, "E": 20,
                             "F": 16, "G": 16, "H": 14, "I": 10})
        self._topo(ws, "GERENCIAMENTO DE CONTRATACOES")

        r = 4
        _header(ws, r, 2, 9,
                "15. Controle de Mapas de Contratacoes",
                COR_AZUL_MEDIO, 10)
        r += 1

        for col, txt in [(2, "#"), (3, "Mes"), (4, "Servico"),
                         (5, "Fornecedor"), (6, "Valor Contratado (R$)"),
                         (7, "Orcamento Atualizado (R$)"),
                         (8, "Desvio (R$)"), (9, "Indice")]:
            _cell(ws, r, col, txt, negrito=True, tam=9,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        for i, m in enumerate(self.dados.get("controle_mapas_contratacoes", [])):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            desvio = m.get("desvio", 0)
            indice = m.get("indice", 1.0)
            bg_d = COR_VERDE_OK if desvio <= 0 else COR_VERMELHO_BG
            bg_i = COR_VERDE_OK if indice >= 0.95 else (
                COR_AMARELO if indice >= 0.85 else COR_VERMELHO_BG)

            _cell(ws, r, 2, m["id"], tam=9, bg=bg, h="center")
            _cell(ws, r, 3, m.get("mes", ""), tam=9, bg=bg, h="center")
            _cell(ws, r, 4, m["servico"], tam=9, bg=bg)
            _cell(ws, r, 5, m["fornecedor"], tam=9, bg=bg)
            _cell(ws, r, 6, m["valor_contratado"], tam=9, bg=bg,
                  h="center", fmt="#,##0.00")
            _cell(ws, r, 7, m["orcamento_atualizado"], tam=9, bg=bg,
                  h="center", fmt="#,##0.00")
            _cell(ws, r, 8, desvio, tam=9, bg=bg_d, h="center",
                  fmt="#,##0.00")
            _cell(ws, r, 9, indice, tam=9, bg=bg_i, h="center",
                  fmt="0.000")

        # Totais
        r += 1
        total_contrat = sum(m["valor_contratado"]
                            for m in self.dados.get("controle_mapas_contratacoes", []))
        total_orc = sum(m["orcamento_atualizado"]
                        for m in self.dados.get("controle_mapas_contratacoes", []))
        total_desv = total_orc - total_contrat
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        _cell(ws, r, 2, "TOTAL", negrito=True, tam=9,
              bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO, h="center")
        _cell(ws, r, 6, total_contrat, negrito=True, tam=9,
              bg=COR_AZUL_CLARO, h="center", fmt="#,##0.00")
        _cell(ws, r, 7, total_orc, negrito=True, tam=9,
              bg=COR_AZUL_CLARO, h="center", fmt="#,##0.00")
        bg_t = COR_VERDE_OK if total_desv <= 0 else COR_VERMELHO_BG
        _cell(ws, r, 8, total_desv, negrito=True, tam=9,
              bg=bg_t, h="center", fmt="#,##0.00")

    # ── 16) CRONOGRAMA DE SUPRIMENTOS ────────────────────────────────────────

    def _criar_cronograma_suprimentos(self):
        ws = self.wb.create_sheet("16) Cronograma de Suprimentos")
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {
            "B": 4, "C": 5, "D": 28, "E": 12, "F": 6, "G": 12,
            "H": 12, "I": 6, "J": 12, "K": 12, "L": 6, "M": 12,
            "N": 12, "O": 6, "P": 12, "Q": 12, "R": 6, "S": 12, "T": 8,
        })
        self._topo(ws, "CRONOGRAMA DE CONTRATACOES - SUPRIMENTOS")

        r = 4
        _header(ws, r, 2, 20,
                "16. Cronograma de Contratacoes (Suprimentos)",
                COR_AZUL_MEDIO, 10)
        r += 1

        for c1, c2, txt in [
            (5, 7, "CARTA CONVITE"),
            (8, 10, "ENTREGA DAS PROPOSTAS"),
            (11, 13, "EQUALIZACAO"),
            (14, 16, "APROVACAO"),
            (17, 19, "CONTRATACAO"),
        ]:
            ws.merge_cells(start_row=r, start_column=c1,
                          end_row=r, end_column=c2)
            _cell(ws, r, c1, txt, negrito=True, tam=8,
                  bg=COR_AZUL_ESCURO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        r += 1
        for col, txt in [
            (2, "#"), (3, "ID"), (4, "Servico"),
            (5, "Previsto"), (6, "Prazo"), (7, "Realizado"),
            (8, "Previsto"), (9, "Prazo"), (10, "Realizado"),
            (11, "Previsto"), (12, "Prazo"), (13, "Realizado"),
            (14, "Previsto"), (15, "Prazo"), (16, "Realizado"),
            (17, "Previsto"), (18, "Prazo"), (19, "Realizado"),
            (20, "Status"),
        ]:
            _cell(ws, r, col, txt, negrito=True, tam=8,
                  bg=COR_AZUL_MEDIO, cor_f=COR_TITULO_TEXTO,
                  h="center", wrap=True)

        for i, sup in enumerate(self.dados.get("cronograma_suprimentos", [])):
            r += 1
            ws.row_dimensions[r].height = 16
            bg = _zebra(i)
            _cell(ws, r, 2, i + 1, tam=8, bg=bg, h="center")
            _cell(ws, r, 3, sup["id"], tam=8, bg=bg, h="center")
            _cell(ws, r, 4, sup["servico"], tam=8, bg=bg)

            campos = [
                sup.get("carta_convite_previsto"),
                sup.get("carta_convite_prazo"),
                sup.get("carta_convite_realizado"),
                sup.get("entrega_propostas_previsto"),
                sup.get("entrega_propostas_prazo"),
                sup.get("entrega_propostas_realizado"),
                sup.get("equalizacao_previsto"),
                sup.get("equalizacao_prazo"),
                sup.get("equalizacao_realizado"),
                sup.get("aprovacao_previsto"),
                sup.get("aprovacao_prazo"),
                sup.get("aprovacao_realizado"),
                sup.get("contratacao_previsto"),
                sup.get("contratacao_prazo"),
                sup.get("contratacao_realizado"),
            ]
            for j, val in enumerate(campos):
                col = j + 5
                is_data = (j % 3 != 1)
                parsed = _parse_data(val) if is_data and val else val
                fmt = "DD/MM/YYYY" if isinstance(parsed, datetime) else None
                _cell(ws, r, col, parsed, tam=8, bg=bg,
                      h="center", fmt=fmt)

            status = sup.get("status", "")
            bg_s = COR_VERDE_OK if status == "OK" else COR_AMARELO
            _cell(ws, r, 20, status, negrito=True, tam=8,
                  bg=bg_s, h="center")

    # ── METODO PRINCIPAL ──────────────────────────────────────────────────────

    def gerar(self, caminho_saida: str):
        etapas = [
            ("Capa",                      self._criar_capa),
            ("Sumario",                   self._criar_sumario),
            ("Resumo e Indicadores",      self._criar_resumo_indicadores),
            ("Fluxo de Caixa",            self._criar_fluxo_caixa),
            ("Analise Financeira",        self._criar_analise_financeira),
            ("Prototipo e Histograma",    self._criar_prototipo_histograma),
            ("Farol de Metas",            self._criar_farol_metas),
            ("Metas Proximo Mes",         self._criar_metas_proximo_mes),
            ("Legalizacao",               self._criar_legalizacao),
            ("Tabela de Aporte",          self._criar_tabela_aporte),
            ("Gerenc. Contratacoes",      self._criar_gerenciamento_contratacoes),
            ("Mapas de Contratacoes",     self._criar_controle_mapas_contratacoes),
            ("Cronograma de Suprimentos", self._criar_cronograma_suprimentos),
        ]

        print("BYM - Gerador de Relatorio Gerencial")
        print(f"Periodo: {self._periodo()}")
        print(f"Projeto: {self.dados['projeto']['nome']}")
        print("-" * 50)

        for nome, func in etapas:
            print(f"  Criando: {nome}...")
            func()

        self.wb.save(caminho_saida)
        print("-" * 50)
        print(f"Relatorio salvo em: {caminho_saida}")

    def gerar_bytes(self, log_fn=None) -> bytes:
        """Gera o workbook em memoria e retorna os bytes do .xlsx."""
        etapas = [
            ("Capa",                      self._criar_capa),
            ("Sumario",                   self._criar_sumario),
            ("Resumo e Indicadores",      self._criar_resumo_indicadores),
            ("Fluxo de Caixa",            self._criar_fluxo_caixa),
            ("Analise Financeira",        self._criar_analise_financeira),
            ("Prototipo e Histograma",    self._criar_prototipo_histograma),
            ("Farol de Metas",            self._criar_farol_metas),
            ("Metas Proximo Mes",         self._criar_metas_proximo_mes),
            ("Legalizacao",               self._criar_legalizacao),
            ("Tabela de Aporte",          self._criar_tabela_aporte),
            ("Gerenc. Contratacoes",      self._criar_gerenciamento_contratacoes),
            ("Mapas de Contratacoes",     self._criar_controle_mapas_contratacoes),
            ("Cronograma de Suprimentos", self._criar_cronograma_suprimentos),
        ]
        for nome, func in etapas:
            if log_fn:
                log_fn(f"Criando: {nome}...")
            func()
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
